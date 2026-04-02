"""
LLM 語意驗證器。

不靠關鍵字比對，讓 LLM 理解整體 log 內容，
判斷步驟是否真正成功——能區分「Python WARNING 不代表失敗」
與「真正的 Exception / 資料異常」。

支援：
- 文字檔讀取前 N 行供 LLM 判斷
- CSV / JSON / Excel 結構化摘要（欄位、行數、樣本）
- 圖片檔以 base64 傳給 Vision model 做視覺驗證
"""
import base64
import csv
import io
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage, SystemMessage

from config import GROQ_API_KEY, GROQ_MODEL_MAIN


@dataclass
class ValidationResult:
    status: str      # "ok" | "warning" | "failed"
    reason: str      # 中文說明
    suggestion: str  # LLM 建議的修復方向（failed 時才有意義）


_llm: Optional[ChatGroq] = None


def _get_llm() -> ChatGroq:
    global _llm
    if _llm is None:
        _llm = ChatGroq(
            api_key=GROQ_API_KEY,
            model=GROQ_MODEL_MAIN,
            temperature=0,
        )
    return _llm


# ── 檔案內容讀取 ──────────────────────────────────────────────────────────────

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}
STRUCTURED_EXTS = {'.csv', '.json', '.jsonl', '.xlsx', '.xls'}
TEXT_EXTS = {'.txt', '.log', '.md', '.html', '.xml', '.yaml', '.yml', '.py', '.sh', '.js', '.ts'}
MAX_TEXT_LINES = 50
MAX_CSV_ROWS = 10


def _read_file_content(path: Optional[str]) -> dict:
    """
    讀取輸出檔案，回傳結構化資訊供 LLM 分析。

    Returns:
        {
            "summary": str,       # 給 prompt 的文字摘要
            "image_b64": str|None # base64 圖片（僅圖檔）
            "image_mime": str|None
        }
    """
    result = {"summary": "", "image_b64": None, "image_mime": None}
    if not path:
        return result

    p = Path(path).expanduser()
    if not p.exists():
        return result

    # 目錄：列出檔案清單
    if p.is_dir():
        files = sorted(p.iterdir())[:20]
        listing = "\n".join(f"  {'📁' if f.is_dir() else '📄'} {f.name} ({f.stat().st_size:,} bytes)" for f in files)
        result["summary"] = f"目錄內容（前 20 項）：\n{listing}"
        return result

    ext = p.suffix.lower()

    # 圖片檔 → base64
    if ext in IMAGE_EXTS:
        try:
            data = p.read_bytes()
            if len(data) <= 20 * 1024 * 1024:  # ≤ 20MB
                result["image_b64"] = base64.b64encode(data).decode()
                mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                           '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
                result["image_mime"] = mime_map.get(ext, 'image/png')
                result["summary"] = f"圖片檔 {p.name}（{len(data):,} bytes），已附圖供視覺分析"
        except Exception as e:
            result["summary"] = f"圖片讀取失敗：{e}"
        return result

    # CSV
    if ext == '.csv':
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                result["summary"] = "CSV 檔案為空"
                return result
            header = rows[0]
            data_rows = rows[1:]
            sample = data_rows[:MAX_CSV_ROWS]
            sample_str = "\n".join([",".join(r) for r in sample])
            result["summary"] = (
                f"CSV 檔案：{p.name}\n"
                f"欄位（{len(header)} 個）：{', '.join(header)}\n"
                f"資料行數：{len(data_rows)}\n"
                f"前 {min(len(sample), MAX_CSV_ROWS)} 行樣本：\n{sample_str}"
            )
        except Exception as e:
            result["summary"] = f"CSV 讀取失敗：{e}"
        return result

    # JSON / JSONL
    if ext in ('.json', '.jsonl'):
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            if ext == '.jsonl':
                lines = [l for l in text.strip().split('\n') if l.strip()]
                result["summary"] = (
                    f"JSONL 檔案：{p.name}，共 {len(lines)} 行\n"
                    f"前 {min(5, len(lines))} 行樣本：\n" +
                    "\n".join(lines[:5])
                )
            else:
                data = json.loads(text)
                if isinstance(data, list):
                    sample = json.dumps(data[:5], ensure_ascii=False, indent=2)
                    result["summary"] = f"JSON 陣列：{p.name}，共 {len(data)} 筆\n前 5 筆樣本：\n{sample}"
                elif isinstance(data, dict):
                    keys = list(data.keys())[:20]
                    result["summary"] = f"JSON 物件：{p.name}\n鍵（前 20 個）：{', '.join(keys)}\n內容預覽：\n{json.dumps(data, ensure_ascii=False, indent=2)[:1000]}"
                else:
                    result["summary"] = f"JSON 檔案：{p.name}\n內容：{text[:500]}"
        except Exception as e:
            result["summary"] = f"JSON 讀取失敗：{e}"
        return result

    # Excel
    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=MAX_CSV_ROWS + 1, values_only=True))
                if not rows:
                    sheets_info.append(f"  Sheet「{sheet_name}」：空")
                    continue
                header = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]
                total_rows = ws.max_row - 1 if ws.max_row else 0
                sample_lines = []
                for r in data_rows[:MAX_CSV_ROWS]:
                    sample_lines.append(",".join(str(c) if c is not None else "" for c in r))
                sheets_info.append(
                    f"  Sheet「{sheet_name}」：{total_rows} 行，{len(header)} 欄\n"
                    f"    欄位：{', '.join(header)}\n"
                    f"    前 {len(sample_lines)} 行：\n    " + "\n    ".join(sample_lines)
                )
            wb.close()
            result["summary"] = f"Excel 檔案：{p.name}，共 {len(wb.sheetnames)} 個 Sheet\n" + "\n".join(sheets_info)
        except ImportError:
            result["summary"] = f"Excel 檔案：{p.name}（需安裝 openpyxl 才能讀取內容）"
        except Exception as e:
            result["summary"] = f"Excel 讀取失敗：{e}"
        return result

    # 一般文字檔（含未知副檔名）
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= MAX_TEXT_LINES:
                    break
                lines.append(line.rstrip())
        total_size = p.stat().st_size
        result["summary"] = (
            f"文字檔：{p.name}（{total_size:,} bytes）\n"
            f"前 {len(lines)} 行：\n" + "\n".join(lines)
        )
    except Exception:
        result["summary"] = f"檔案 {p.name} 無法以文字方式讀取"

    return result


# ── 主驗證函式 ─────────────────────────────────────────────────────────────────

async def validate_step(
    step_name: str,
    command: str,
    exit_code: int,
    stdout: str,
    stderr: str,
    output_path: Optional[str],
    output_expect: Optional[str],
    logger: logging.Logger,
) -> ValidationResult:
    """
    使用 LLM 語意分析執行結果，回傳結構化驗證結論。

    LLM 會考量：
    - exit code 與其含意
    - stdout/stderr 的語意（區分警告與錯誤）
    - 輸出檔案是否存在、大小是否合理
    - 輸出檔案內容（文字前 50 行 / CSV 結構 / Excel 摘要）
    - 圖片檔案以視覺方式驗證
    - 是否符合 expect 描述的期望
    """
    # 收集輸出檔案資訊
    file_info = _check_output_file(output_path)
    file_content = _read_file_content(output_path)

    # 截取重要片段（節省 token）
    stdout_tail = stdout[-1000:] if len(stdout) > 1000 else stdout
    stderr_tail = stderr[-500:] if len(stderr) > 500 else stderr

    prompt_text = f"""你是一個精確的 pipeline 步驟驗證器。
分析以下執行結果，判斷步驟是否成功。

【步驟資訊】
名稱：{step_name}
命令：{command}
Exit Code：{exit_code}
預期輸出描述：{output_expect or "無特定要求"}
輸出路徑：{output_path or "無"}
檔案狀態：{file_info}

【stdout（最後部分）】
```
{stdout_tail or "（無輸出）"}
```

【stderr（最後部分）】
```
{stderr_tail or "（無輸出）"}
```"""

    # 加入檔案內容摘要
    if file_content["summary"]:
        prompt_text += f"""

【輸出檔案內容】
{file_content["summary"]}"""

    # 如果是圖片，加入視覺分析提示
    if file_content["image_b64"]:
        prompt_text += """

【圖片分析】
已附上輸出的圖片檔案，請以視覺方式分析圖片內容是否符合預期描述。
檢查圖片是否正常渲染、內容是否完整、是否符合期望。"""

    prompt_text += """

請只回傳以下 JSON，不要加任何其他文字：
{
  "status": "ok",
  "reason": "一句話說明判斷結果",
  "suggestion": "如果 failed，給出修復建議；ok 時留空字串"
}

【判斷規則】
- "ok"：步驟成功，exit code 0，輸出符合預期（若有）
- "warning"：步驟完成但有非致命問題（如 deprecation warning、部分資料遺失），建議人工確認
- "failed"：步驟失敗，需要介入（exit code 非 0 且 stderr 有真實錯誤、Exception、缺少必要輸出檔案等）

注意：Python DeprecationWarning、UserWarning 不代表失敗；只有真正的 Exception / Error / 致命問題才判為 failed。"""

    try:
        llm = _get_llm()

        # 構建 message content（支援圖片 vision）
        if file_content["image_b64"]:
            content = [
                {"type": "text", "text": prompt_text},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{file_content['image_mime']};base64,{file_content['image_b64']}"
                    },
                },
            ]
        else:
            content = prompt_text

        response = await llm.ainvoke([
            SystemMessage(content="你是一個精確的 pipeline 驗證器，只回傳 JSON 格式。"),
            HumanMessage(content=content),
        ])

        raw = response.content.strip()
        # 去除 markdown code block（如果有）
        if "```" in raw:
            parts = raw.split("```")
            raw = parts[1].strip()
            if raw.startswith("json"):
                raw = raw[4:].strip()

        data = json.loads(raw)
        result = ValidationResult(
            status=data.get("status", "failed"),
            reason=data.get("reason", ""),
            suggestion=data.get("suggestion", ""),
        )
        logger.info(f"[{step_name}] 驗證：{result.status} — {result.reason}")
        return result

    except Exception as e:
        logger.error(f"[{step_name}] LLM 驗證失敗：{e}，退回 exit code 判斷")
        # Fallback：純 exit code 判斷
        if exit_code == 0:
            return ValidationResult(
                status="ok",
                reason=f"Exit code 0（LLM 驗證服務暫時不可用：{e}）",
                suggestion="",
            )
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exit_code}（LLM 驗證服務暫時不可用：{e}）",
            suggestion="請檢查 log 檔取得詳細錯誤訊息",
        )


def _check_output_file(path: Optional[str]) -> str:
    """取得輸出檔案或目錄的基本資訊"""
    if not path:
        return "無需檢查"
    p = Path(path).expanduser()
    if not p.exists():
        return "❌ 路徑不存在"
    if p.is_dir():
        files = list(p.iterdir())
        if not files:
            return "⚠ 目錄存在但為空"
        total = sum(f.stat().st_size for f in files if f.is_file())
        return f"✅ 目錄存在，共 {len(files)} 個檔案，總大小：{total:,} bytes"
    size = p.stat().st_size
    if size == 0:
        return "⚠ 檔案存在但為空（0 bytes）"
    return f"✅ 檔案存在，大小：{size:,} bytes"
